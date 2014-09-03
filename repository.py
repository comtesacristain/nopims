#!/usr/bin/python
ACTIVITY_NAME_COL = 3
TITLE_COL = 4
STAGED_COL = 9
COPIED_COL = 8
NOPTA_FILE = "/nas/energy/ideas/RDIS/NOPIMS_repository_remediation/msexton1/NOPTA_20120101_20140728_OpenFile_Well_List.xlsx"
NOPTA_SHEET_NAME = "NOPTA-OF-Wells"
WELLS_ROOT = "/nas/pmd/repos/open/Wells/Regulated"
DEST_DIR = "/nas/energy/ideas/RDIS/NOPIMS_repository_remediation/msexton1"

associated_wells = []



import os, re, xlrd, xlwt, shutil

import openpyxl as px

def run_me():
    global associated_wells
    workbook = xlrd.open_workbook(NOPTA_FILE)
    sheet = workbook.sheet_by_name(NOPTA_SHEET_NAME)
    num_rows = sheet.nrows
    i = 1
    while i < num_rows:    
        row = sheet.row(i)
        title = row[TITLE_COL].value
        activity_name = row[ACTIVITY_NAME_COL].value
        staged = row[STAGED_COL].value
        copied = row[COPIED_COL].value
        if not staged and not copied:
            state_folder = title.split('/')[0].split('-')[0]
            state_folder = re.sub("[0-9]",'',state_folder)
            folder_key= folder_key_parse(activity_name)
            if os.path.isdir(os.path.join(WELLS_ROOT,state_folder)):
                search_path = os.path.join(WELLS_ROOT,state_folder)
                paths=find_paths(search_path,folder_key)
            else:
                search_path = WELLS_ROOT
                paths=find_paths(search_path,folder_key,2)
                associate_wells_and_paths(title,activity_name,paths)
            #associate_wells_and_paths(title,activity_name,paths)
        i += 1    
    print associated_wells.__len__()
    copy_wells()
    
def copy_wells():
    global associated_wells
    for aw in associated_wells:
        for a in aw["activities"]:
            for p in a["paths"]:
                os.system("cp -r {0} {1}".format(p,DEST_DIR))
                mark_copied(a["name"])
    os.system("chmod -R 775 {0}".format(DEST_DIR))

def mark_copied(n):
    wb = px.load_workbook(NOPTA_FILE,use_iterators=True)
    ws = wb.get_sheet_by_name(NOPTA_SHEET_NAME)
    for i, row in enumerate(ws.iter_rows()):
        if row[ACTIVITY_NAME_COL] == n:
            ws.cell(row = i+1,column = COPIED_COL+1).value = 'X'
    wb.save(NOPTA_FILE)
        
    
    

def associate_wells_and_paths(t,an,p):
    global associated_wells
    print associated_wells
    title = None
    for aw in associated_wells:
        if aw["title"] == t:
            title = associated_wells.pop(associated_wells.index(aw))
    if title is None:
        title = {"title": t, "activities": []}
    activity = None
    for a in title["activities"]:
        if a["name"] == an:
            activity = title["activities"].pop(title["activities"].index(a))
    if activity is None:
        activity = {"name": an, "paths": []}
    activity["paths"] = list(set(activity["paths"] + p)) 
    title["activities"].append(activity)
    associated_wells.append(title)
    

def find_paths(sp, ak, lvl = 1):
    paths,directories = [],[]
    dirs=os.listdir(sp)
    if lvl > 1:
        for x in dirs:
            directories += find_paths(os.path.join(sp,x),ak, lvl-1)
    directories += filter(lambda x: ak in x, dirs)
    paths = map(lambda x: os.path.join(sp,x), directories)
    return paths


def folder_key_parse(an):
    key = re.sub("\([^)]+\)",'',an).strip() # Remove parentheses and strip trailing and leading spaces
    key =  re.sub("(ST([0-9])|L([0-9]))",'',key).strip() #Remove SH
    return key.replace(' ','_')



run_me()
    
